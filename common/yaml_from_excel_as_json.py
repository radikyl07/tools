import openpyxl, yaml
from openpyxl.utils.cell import get_column_letter 
from json_csv_key_to_list import json_csv_key_to_list
from json_remove_none_data import json_remove_none_data


#### Read ';;' separated yaml data in cells from a designated column in a specified sheet from an 
#       Excel workbook, and return python list (JSON) data with each Excel cell as a list item.
#       Also, expands key names with suffix __csv into list items
#       Also, removes None type values (list item or dict val) from data
def yaml_from_excel_as_json(
    wbn,                 # workbook name
    wsn,                 # worksheet name
    col_yaml_name,      # column header containing yaml code i.e. _yaml
    ):
    # open workbook
    try:
        wb = openpyxl.load_workbook(
            wbn,
            data_only=True,
            )
    except Exception as e:
        print(f"ERROR - reading file {wbn}. Exception: {e}")
        exit()

    # get worksheet
    if wsn in wb.sheetnames:
        ws = wb[wsn]
    else:
        print(f"ERROR - Unable to open sheet {wsn} in workbook {wbn}")
        exit()

    # get header row
    ws_header = [cell.value for cell in ws["1"]]

    # get entire yaml data column contents, excludes header label
    try:
        col_yaml = get_column_letter(ws_header.index(col_yaml_name) + 1)
    except Exception as e:
        print(f"ERROR - Unable to get column {col_yaml_name}. Exception: {e}")
        exit()
    col_yaml_data = [cell.value for i,cell in enumerate(ws[col_yaml]) if i>0]
    
    # replace placeholder newline delimeter ';;' with actual newline '\n', if cell is not None
    yaml_data_str = "".join(
        list(
            map(
                lambda x: x.replace(";;", "\n") if x is not None else "", 
                col_yaml_data,
                )
            )
        )
    
    # convert yaml string to python list / json object
    try:
        json_data = yaml.load(yaml_data_str, Loader=yaml.FullLoader)
    except Exception as e:
        print(f"ERROR - in parsing yaml data. Exception: {e}")
        exit()
    
    # expand the value (a csv) of any key name ending in '__csv' into a list and rename key
    # removing the '__csv' component from it
    jd2 = json_csv_key_to_list(json_data, '__csv')
    
    # remove None type values from the data
    jd3 = json_remove_none_data(jd2)
    
    # return json object and original yaml data as read from excel file
    return jd3


#### Example Run - when this file is run directly
if __name__ == "__main__":
    import pprint

    input_data = yaml_from_excel_as_json(
        wbn="yaml_from_excel_as_json.xlsx",
        wsn="data",
        col_yaml_name="_yaml",
        )
    
    pprint.pprint(input_data, indent=4)
