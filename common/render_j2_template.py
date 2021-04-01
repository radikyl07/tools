import json, yaml
from jinja2 import Environment

#### Renders a Jinja2 source template (string) along with a list of dictionary object to
#       generate a list of rendered data (string).

def render_j2_template(
    source,             # source Jinja2 template as a string
    data,               # list of dictionary
    type="text",        # type of 'source' template [text (default), json, yaml]
    validate=False,     # validate rendered output syntax? valid for type [json, yaml]
    minify=False,       # minify json. implicit validation. valid for type [json].
    ):
    env = Environment()
    env.trim_blocks = True
    env.lstrip_blocks = True
    env.rstrip_blocks = True
    
    template = env.from_string(source)

    rendered = []
    for td in data:
        tr = template.render(td)
        
        try:
            if type in ['json', 'yaml']:
                if (type == 'json' and validate) or (type == 'json' and minify):
                    jtr = json.loads(tr)
                    if minify:
                        tr = json.dumps(jtr, separators=(',', ':'))
                        
                if type == 'yaml' and validate == True:
                    yaml.load(tr, Loader=yaml.FullLoader)
                    
        except Exception:
            print(f"="*32)
            print(f"WARNING - '{type}' data validation failed. Skipping...")
            print(f"{tr}")
            print(f"-"*32)
            continue
        
        rendered.append(tr)
        
    return rendered


#### Example Run - when this file is run directly
if __name__ == "__main__":
    # import modules
    import openpyxl
    
    # INPUT VARIABLES
    wbn = "render_j2_template.xlsx"
    wsn = "data1"
    
    # read template source file
    with open("render_j2_template.src.txt") as f:
        template_source = f.read()

    # read template data file
    template_data = []
    
    wb = openpyxl.load_workbook(
        wbn,
        data_only=True,
        )
    ws = wb[wsn]
    
    header = [cell.value for cell in ws[1]]
    for row in ws[2: ws.max_row]:
        values = {}
        for key, cell in zip(header, row):
            values[key] = cell.value
        template_data.append(values)
        
    # render template - checks rendered template is valid Json and minifies it.
    rendered = render_j2_template(
        template_source, 
        template_data,
        type='json',
        minify=True,
    )
    
    # print original template and rendered data
    print(f"*** TEMPLATE (str) ***\n{template_source}\n")
    print(f"*** RENDERED (str) ***")
    for r in rendered:
        print(f"{r}")

