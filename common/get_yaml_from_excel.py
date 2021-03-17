import openpyxl, yaml
from openpyxl.utils.cell import get_column_letter 


#### Read ';;' separated yaml data in cells from a designated column in a specified sheet from an 
#       Excel workbook, and return python list (JSON) data with each Excel cell as a list item.
def get_yaml_from_excel(
    wbn,                 # workbook name
    wsn,                 # worksheet name
    col_yaml_name,      # column header containing yaml code i.e. _yaml
    ):
    # open workbook
    wb = openpyxl.load_workbook(
        wbn,
        data_only=True,
        )

    # get worksheet
    if wsn in wb.sheetnames:
        ws = wb[wsn]
    else:
        print(f"ERROR - Unable to open sheet {ws} in workbook {wb}")
        exit()

    # get header row
    ws_header = [cell.value for cell in ws["1"]]

    # get entire yaml data column contents, excluding header row
    col_yaml = get_column_letter(ws_header.index(col_yaml_name) + 1)
    col_yaml_data = [cell.value for i,cell in enumerate(ws[col_yaml]) if i>0]
    
    # replace placeholder newline delimeter ';;' with actual newline '\n'
    yaml_data_str = "".join(list(map(lambda x: x.replace(";;", "\n"), col_yaml_data)))
    
    # convert yaml string to python list / json object
    yaml_data_list = yaml.load(yaml_data_str, Loader=yaml.FullLoader)
    
    # return yaml data in string format. 
    # - Calling func can do 'yaml.load(yaml_data, Loader=yaml.FullLoader)' to get a list obj
    return yaml_data_list
