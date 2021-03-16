import requests
import json, yaml
import pprint
import openpyxl
from openpyxl.utils.cell import get_column_letter
from decouple import config



#### Initialize

AMP_BASE_URL = 'https://api.apjc.amp.cisco.com/v1/'
AMP_CLIENT_ID = config('AMP_CLIENT_ID')
AMP_API_KEY = config('AMP_API_KEY')


 
#### Read YAML data from Excel worksheet

excel_wb = "amp_create_groups.xlsx"
excel_ws = "groups"
excel_header_yaml = "_yaml"

# open workboot
wb = openpyxl.load_workbook(
    excel_wb,
    data_only=True,
    )

# get worksheet
if excel_ws in wb.sheetnames:
    ws = wb[excel_ws]
else:
    print(f"ERROR - Unable to open sheet {excel_ws} in workbook {excel_wb}")
    exit()

# get header row
ws_header = [cell.value for cell in ws["1"]]

# get entire yaml data column
col_yaml = get_column_letter(ws_header.index(excel_header_yaml) + 1)
col_yaml_data = [cell.value for i,cell in enumerate(ws[col_yaml]) if i>0]
yaml_data = "".join(list(map(lambda x: x.replace(";", "\n"), col_yaml_data)))



#### Create AMP Groups

input_data = yaml.load(yaml_data, Loader=yaml.FullLoader)
headers={
    "content-type": "application/json",
    "accept": "application/json",
}

for k,v in enumerate(input_data):
    data = json.dumps(v)
    resp = requests.post(
        url=AMP_BASE_URL+"groups",
        headers=headers,
        auth=(AMP_CLIENT_ID, AMP_API_KEY),
        data=data,
    )
    
    if k==0:
        print(f"API Rate Limit Remaining/Total/Seconds-to-reset: "+
              f"{resp.headers['X-RateLimit-Remaining']}/"+
              f"{resp.headers['X-RateLimit-Limit']}/"+
              f"{resp.headers['X-RateLimit-Reset']}")
    
    if resp.status_code == 201:
        print(f"OK - {v['name']} created. GUID = {resp.json()['data']['guid']}")
    else:
        e = resp.json()["errors"][0]
        print(f"FAIL - {v['name']} failed to create."+
              f" API Result = {resp.status_code} {resp.reason}."+
              f" ERROR = {e['error_code']} - {e['description']} - {e['details'][0]}"
              )

